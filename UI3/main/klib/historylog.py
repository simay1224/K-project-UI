import pandas as pd
import os.path
from openpyxl import load_workbook
import pdb

class Historylog(object):
    """ save the user's data from furture analysis and comparision.
        different exercises will be record in different sheets
        so far, we have 7 exercises as following.
        #1 : Muscle Tighting Deep Breathing
        #2 : Over The Head Pumping
        #3 : Push Down Pumping
        #4 : Horizontal Pumping
        #5 : Reach to the Sky  ############THIS WILL BE HIDDEN
        #6 : Shoulder Rolls
        #7 : Clasp and Spread
    """

    def __init__(self):
        # HIDING THE EXERCISE 5: reach to sky
        #self.excelPath = "./output/log.xlsx"
        self.excelPath = "./output/log_without_exercise5_reach_to_sky.xlsx"
        self.order = [0,1,2,2,3,4,5]
        #self.exercises = ["",
        #                    "1. Muscle Tighting Deep Breathi",
        #                    "2. Over The Head Pumping",
        #                    "3. Push Down Pumping",
        #                    "4. Horizontal Pumping",
        #                    "5. Reach to the Sky",
        #                    "6. Shoulder Rolls",
        #                    "7. Clasp and Spread"]

        self.exercises = ["",
                            "1. Muscle Tighting Deep Breathi",
                            "2. Over The Head Pumping",
                            "3. Push Down Pumping",
                            "4. Horizontal Pumping",
                            "5. Shoulder Rolls",
                            "6. Clasp and Spread"]
        # record feature
        common = ["name", "age", "gender", "time"]
        backup = ["errmsg"]
        # exercise recoding feature
        brth  = ["Min depth (in mm)", "Max depth (in mm)", "Average depth (in mm)"]  # breathing exercise feature
        hs    = ["Sync rate"]                         # hand exercise feature => hand breath sync rate
        shld  = ["Max shoulder depth (in mm)", "Min shoulder depth (in mm)"]
        clsp  = ["Max hold time", "Min hold time", "Average hold time"]
        swing = ["Max right bending angle", "Min right bending angle",
                 "Max left bending angle", "Min left bending angle"]
        exer3 = ['1st lower right elbow angle', '1st straight right elbow angle',
                 '2nd lower right elbow angle', '2nd straight right elbow angle',
                 '3rd lower right elbow angle', '3rd straight right elbow angle',
                 '4th lower right elbow angle', '4th straight right elbow angle',
                 'Average right hand push down angle', 'Average right hand straight angle',
                 '1st lower left elbow angle', '1st straight left elbow angle',
                 '2nd lower left elbow angle', '2nd straight left elbow angle',
                 '3rd lower left elbow angle', '3rd straight left elbow angle',
                 '4th lower left elbow angle', '4th straight left elbow angle',
                 'Average left hand push down angle', 'Average left hand straight angle']
        exer4 = ['1st right arm horizontal angle (H-close)', '1st right arm horizontal angle (T-pose)',
                 '2nd right arm horizontal angle (H-close)', '2nd right arm horizontal angle (T-pose)',
                 '3rd right arm horizontal angle (H-close)', '3rd right arm horizontal angle (T-pose)',
                 '4th right arm horizontal angle (H-close)', '4th right arm horizontal angle (T-pose)',
                 'Average right arm horizontal angle (H-close)', 'Average right arm horizontal angle (T-pose)',
                 '1st left arm horizontal angle (H-close)', '1st left arm horizontal angle (T-pose)',
                 '2nd left arm horizontal angle (H-close)', '2nd left arm horizontal angle (T-pose)',
                 '3rd left arm horizontal angle (H-close)', '3rd left arm horizontal angle (T-pose)',
                 '4th left arm horizontal angle (H-close)', '4th left arm horizontal angle (T-pose)',
                 'Average left arm horizontal angle (H-close)', 'Average left arm horizontal angle (T-pose)']
        # ideal value
        icommon = ['$IDEAL VALUE$', 'NaN', 'NaN', 'NaN']
        ibrth = ['bigger is better', 'bigger is better', 'bigger is better']
        ihs   = [100]
        ishld = ['bigger is better', 'bigger is better']
        iclsp  = ['bigger is better', 'bigger is better', 'bigger is better']
        iswing = [90, 'bigger is better', 90, 'bigger is better']
        iexer3 = [10, 180, 10, 180, 10, 180, 10, 180, 10, 180,
                    10, 180, 10, 180, 10, 180, 10, 180, 10, 180]
        iexer4 = [90]*20
        ibackup = []
        ibackup = []
        # cols title for different exercise sheets
        self.colname = {}
        self.colname[1] = common + brth + backup
        self.colname[2] = common + brth + hs + backup
        self.colname[3] = common + exer3 + backup
        self.colname[4] = common + exer4 + backup
        #self.colname[5] = common + swing + backup  # hide reach to sky
        #self.colname[6] = common + shld + backup
        #self.colname[7] = common + clsp + backup
        
        self.colname[5] = common + shld + backup
        self.colname[6] = common + clsp + backup

        # ideal number
        self.icol = {}
        self.icol[1] = icommon + ibrth + ibackup
        self.icol[2] = icommon + ibrth + ihs + ibackup
        self.icol[3] = icommon + iexer3 + ibackup
        self.icol[4] = icommon + iexer4 + ibackup
        #self.icol[5] = icommon + iswing + ibackup  # hide reach to sky
        #self.icol[6] = icommon + ishld + ibackup
        #self.icol[7] = icommon + iclsp + ibackup

        self.icol[5] = icommon + ishld + ibackup
        self.icol[6] = icommon + iclsp + ibackup

    def newlog(self, sheetnum=7):
        """ if there is not log.xlsx exist, create one with sheetnum sheets
            sheetnum depends on howmany exercise we have
        """
        excelWriter = pd.ExcelWriter(self.excelPath, engine='openpyxl')  #create excel file
        for i in xrange(1, sheetnum+1):
            dataframe = pd.DataFrame(columns = self.colname[i])
            # ideal = pd.DataFrame(columns = self.icol[i])
            if i == 1:
                dataframe.to_excel(excelWriter, 'exercise %s' %i, index=None)
                excelWriter.save()
            else:  # add sheet
                book = load_workbook(excelWriter.path)
                excelWriter.book = book
                dataframe.to_excel(excelWriter, 'exercise %s' %i, index=None)
                excelWriter.save()
            excelWriter.close()
        # append ideal value
        for i in xrange(1, sheetnum+1):
            book = load_workbook(self.excelPath)
            sheet = book['exercise %s' %i]
            sheet.append(self.icol[i])
            book.save(self.excelPath)

    def addsheet(self, num=1):
        """ adding num new sheet
        """
        excelWriter = pd.ExcelWriter(self.excelPath,engine='openpyxl')
        book = load_workbook(excelWriter.path)
        excelWriter.book = book
        for i in range(num):
            sheetnum = len(book.worksheets)
            dataframe.to_excel(excelWriter, 'exercise %s' %sheetnum+1, index=None)
        excelWriter.save()
        excelWriter.close()

    def writein(self, info, exeno, time, data=[], errmsgs=[]):
        """ write the user record in to log file
            info : the infomation user input in the initial msgbox
            exeno : exercise number
            data : list-like object, which contains vary results depend on the exercise type
            time :
        """
        if not os.path.isfile(self.excelPath):
            self.newlog()
        errmsg = errmsgs[self.order[exeno-1]]
        book = load_workbook(self.excelPath)
        sheet = book[self.exercises[exeno]]
        date = '-'.join(map(str,[time.year,time.month,time.day,time.hour,time.minute]))
        userrecord = [info.name, info.id, info.gender, date] + data + errmsg
        sheet.append(userrecord)
        book.save(self.excelPath)
