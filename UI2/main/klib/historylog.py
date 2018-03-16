import pandas as pd
import os.path,pdb
from openpyxl import load_workbook


class Historylog(object):
    """ save the user's data from furture analysis and comparision.
        different exercises will be record in different sheets
        so far, we have 7 exercises as following.
        #1 : Muscle Tighting Deep Breathing
        #2 : Over The Head Pumping
        #3 : Push Down Pumping
        #4 : Horizontal Pumping
        #5 : Reach to the Sky
        #6 : Shoulder Rolls
        #7 : Clasp and Spread
    """

    def __init__(self):
        self.excelPath = "./output/log.xlsx"
        self.order = [0,1,2,2,3,4,5]
        # record feature
        common = ["name", "age", "gender", "time"]
        backup = ["errmsg"]
        # exercise recoding feature
        brth  = ["mindepth", "maxdepth", "avgdepth"]  # breathing exercise feature
        hs    = ["sync_rate"]                         # hand exercise feature => hand breath sync rate
        shld  = ["Max rotation depth", "Min rotation depth"]
        clsp  = ["Max hold time", "Min hold time", "average hold time",
                 "clasp rate"]
        swing = ["Max right bending angle", "Min right bending angle",
                 "Max left bending angle", "Min left bending angle"]
        exer3 = ['1st right push down lower enough?', '1st lower right elbow angle',
                 '1st right hand up straight?', '1st straight right elbow angle',
                 '2nd right push down lower enough?', '2nd lower right elbow angle',
                 '2nd right hand up straight?', '2nd straight right elbow angle',
                 '3rd right push down lower enough?', '3rd lower right elbow angle',
                 '3rd right hand up straight?', '3rd straight right elbow angle',
                 '4th right push down lower enough?','4th lower right elbow angle',
                 '4th right hand up straight?', '4th straight right elbow angle',
                 'average right hand straight angle', 'average right hand push down angle',
                 '1st left push down lower enough?', '1st lower left elbow angle',
                 '1st left hand up straight?', '1st straight left elbow angle',
                 '2nd left push down lower enough?', '2nd lower left elbow angle',
                 '2nd left hand up straight?', '2nd straight left elbow angle',
                 '3rd left push down lower enough?', '3rd lower left elbow angle',
                 '3rd left hand up straight?', '3rd straight left elbow angle',
                 '4th left push down lower enough?','4th lower left elbow angle',
                 '4th left hand up straight?', '4th straight left elbow angle',
                 'average left hand straight angle', 'average left hand push down angle',]

        # cols title for different exercise sheets
        self.colname = {}
        self.colname[1] = common + brth + backup
        self.colname[2] = common + brth + hs + backup
        self.colname[3] = common + exer3 + backup
        self.colname[4] = common + backup
        self.colname[5] = common + swing + backup
        self.colname[6] = common + shld + backup
        self.colname[7] = common + clsp + backup
    
    def newlog(self, sheetnum=7):
        """ if there is not log.xlsx exist, create one with sheetnum sheets
            sheetnum depends on howmany exercise we have
        """
        excelWriter = pd.ExcelWriter(self.excelPath, engine='openpyxl')  #create excel file
        for i in xrange(1, sheetnum+1):
            dataframe = pd.DataFrame(columns = self.colname[i])
            if i == 1:
                dataframe.to_excel(excelWriter, 'exercise %s' %i, index=None)
                excelWriter.save()
            else:  # add sheet
                book = load_workbook(excelWriter.path)
                excelWriter.book = book
                dataframe.to_excel(excelWriter, 'exercise %s' %i, index=None)
                excelWriter.save()
            excelWriter.close()

    def addsheet(self, num=1):
        """ adding num new sheet
        """
        excelWriter = pd.ExcelWriter(self.excelPath,engine='openpyxl')
        book = load_workbook(excelWriter.path)
        excelWriter.book = book
        for i in range(num):
            sheetnum = len(book.worksheets)
            dataframe.to_excel(excelWriter, 'exercise %s' %sheetnum+1, index=None)
        excelWriter.save()
        excelWriter.close()        

    def writein(self, userinfo, exeno, time, data=[], errmsgs=[]):
        """ write the user record in to log file
            userinfo : the infomation user input in the initial msgbox
            exeno : exercise number
            data : list-like object, which contains vary results depend on the exercise type
            time :  
        """
        pdb.set_trace()
        if not os.path.isfile(self.excelPath):
            self.newlog()
        errmsg = errmsgs[self.order[exeno-1]]
        book = load_workbook(self.excelPath)
        sheet = book['exercise %s' %exeno]
        date = '-'.join(map(str,[time.year,time.month,time.day,time.hour,time.minute]))
        userrecord = [userinfo.name, userinfo.age, userinfo.gender, date] + data + errmsg
        sheet.append(userrecord)
        book.save(self.excelPath)


            


